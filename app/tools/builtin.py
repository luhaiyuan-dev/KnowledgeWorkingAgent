import ast
import json
import math
import operator
import os
from typing import Any

import requests
from langchain_core.tools import BaseTool, StructuredTool
from openpyxl import load_workbook
from pydantic import BaseModel, Field

from app.core.config import Settings
from app.rag.pipeline import RagPipeline
from app.security.guardrails import SecurityGuardrails


class KnowledgeSearchInput(BaseModel):
    query: str = Field(description="要在企业知识库中检索的问题")


class FileReadInput(BaseModel):
    file_name: str = Field(description="知识库目录中的文件名")
    max_characters: int = Field(default=6000, ge=100, le=20000)


class CalculatorInput(BaseModel):
    expression: str = Field(description="只包含数字、括号和常见算术运算符的表达式")


class DataQueryInput(BaseModel):
    file_name: str = Field(description="Excel 文件名")
    sheet_name: str | None = Field(default=None, description="可选工作表名称")
    contains: str | None = Field(default=None, description="可选的行文本包含条件")
    limit: int = Field(default=20, ge=1, le=100)


class WebSearchInput(BaseModel):
    query: str = Field(description="需要联网搜索的问题")
    max_results: int = Field(default=5, ge=1, le=10)


class DocumentAnalysisInput(BaseModel):
    file_name: str = Field(description="要分析的知识库文件名")
    question: str = Field(description="分析目标或关注问题")


class SafeCalculator:
    BINARY_OPERATORS = {
        ast.Add: operator.add,
        ast.Sub: operator.sub,
        ast.Mult: operator.mul,
        ast.Div: operator.truediv,
        ast.FloorDiv: operator.floordiv,
        ast.Mod: operator.mod,
        ast.Pow: operator.pow,
    }
    UNARY_OPERATORS = {ast.UAdd: operator.pos, ast.USub: operator.neg}

    def calculate(self, expression: str) -> float | int:
        if len(expression) > 200:
            raise ValueError("表达式过长")
        parsed = ast.parse(expression, mode="eval")
        result = self._evaluate_node(parsed.body)
        if not math.isfinite(float(result)):
            raise ValueError("计算结果不是有限数字")
        return result

    def _evaluate_node(self, node: ast.AST) -> float | int:
        if isinstance(node, ast.Constant) and type(node.value) in {int, float}:
            return node.value
        if isinstance(node, ast.BinOp) and type(node.op) in self.BINARY_OPERATORS:
            left_value = self._evaluate_node(node.left)
            right_value = self._evaluate_node(node.right)
            if isinstance(node.op, ast.Pow) and abs(float(right_value)) > 100:
                raise ValueError("指数绝对值不能超过 100")
            return self.BINARY_OPERATORS[type(node.op)](left_value, right_value)
        if isinstance(node, ast.UnaryOp) and type(node.op) in self.UNARY_OPERATORS:
            return self.UNARY_OPERATORS[type(node.op)](self._evaluate_node(node.operand))
        raise ValueError("表达式包含不允许的语法")


def build_builtin_tools(
    rag: RagPipeline,
    settings: Settings,
    allowed_scopes: set[str],
) -> list[BaseTool]:
    knowledge_root = settings.project_path(settings.rag.knowledge_base_dir)
    guardrails = SecurityGuardrails(settings.app.max_input_chars)
    calculator_engine = SafeCalculator()

    def knowledge_base_search(query: str) -> dict[str, Any]:
        chunks = rag.retrieve(query, allowed_scopes)
        context, citations = rag.build_context_and_citations(chunks, query=query)
        return {
            "context": context,
            "citations": [citation.model_dump() for citation in citations],
        }

    def file_reader(file_name: str, max_characters: int = 6000) -> dict[str, Any]:
        documents = rag.load_named_file(file_name)
        text = "\n\n".join(document.page_content for document in documents)
        return {
            "file_name": file_name,
            "content": text[:max_characters],
            "truncated": len(text) > max_characters,
        }

    def calculator(expression: str) -> dict[str, Any]:
        result = calculator_engine.calculate(expression)
        return {"expression": expression, "result": result}

    def data_query(
        file_name: str,
        sheet_name: str | None = None,
        contains: str | None = None,
        limit: int = 20,
    ) -> dict[str, Any]:
        safe_path = guardrails.resolve_allowed_path(knowledge_root, file_name)
        if safe_path.suffix.lower() != ".xlsx":
            raise ValueError("入门版数据查询工具只读取 .xlsx 文件")
        workbook = load_workbook(safe_path, read_only=True, data_only=True)
        selected_sheet = sheet_name or workbook.sheetnames[0]
        if selected_sheet not in workbook.sheetnames:
            workbook.close()
            raise ValueError(f"工作表不存在：{selected_sheet}")
        worksheet = workbook[selected_sheet]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            workbook.close()
            return {"file_name": file_name, "sheet": selected_sheet, "rows": []}
        headers = [str(value or f"列{index + 1}") for index, value in enumerate(rows[0])]
        result_rows: list[dict[str, Any]] = []
        for row in rows[1:]:
            record = {header: value for header, value in zip(headers, row, strict=False)}
            serialized_record = json.dumps(record, ensure_ascii=False, default=str)
            if contains and contains.lower() not in serialized_record.lower():
                continue
            result_rows.append(record)
            if len(result_rows) >= limit:
                break
        workbook.close()
        return {"file_name": file_name, "sheet": selected_sheet, "rows": result_rows}

    def web_search(query: str, max_results: int = 5) -> dict[str, Any]:
        api_key = os.getenv("TAVILY_API_KEY")
        if not api_key:
            return {
                "status": "configuration_required",
                "message": "未配置 TAVILY_API_KEY，Web 搜索未执行。",
                "results": [],
            }
        response = requests.post(
            "https://api.tavily.com/search",
            json={"api_key": api_key, "query": query, "max_results": max_results},
            timeout=settings.web_search.timeout_seconds,
        )
        response.raise_for_status()
        payload = response.json()
        results = [
            {
                "title": item.get("title"),
                "url": item.get("url"),
                "content": item.get("content"),
            }
            for item in payload.get("results", [])
        ]
        return {"status": "ok", "results": results}

    def document_analysis(file_name: str, question: str) -> dict[str, Any]:
        documents = rag.load_named_file(file_name)
        complete_text = "\n".join(document.page_content for document in documents)
        keywords = [word for word in question.replace("？", " ").split() if len(word) >= 2]
        paragraphs = [line.strip() for line in complete_text.splitlines() if line.strip()]
        matched_paragraphs = [
            paragraph
            for paragraph in paragraphs
            if not keywords or any(keyword in paragraph for keyword in keywords)
        ][:10]
        headings = [
            paragraph
            for paragraph in paragraphs
            if paragraph.startswith("#")
            or paragraph[:2] in {"一、", "二、", "三、", "四、", "五、"}
        ][:20]
        return {
            "file_name": file_name,
            "characters": len(complete_text),
            "headings": headings,
            "relevant_paragraphs": matched_paragraphs,
        }

    return [
        StructuredTool.from_function(
            func=knowledge_base_search,
            name="knowledge_base_search",
            description="搜索企业内部知识库，返回相关片段及来源引用。",
            args_schema=KnowledgeSearchInput,
        ),
        StructuredTool.from_function(
            func=file_reader,
            name="file_reader",
            description="读取知识库允许目录中的 PDF、Word、Excel、TXT 或 Markdown 文件。",
            args_schema=FileReadInput,
        ),
        StructuredTool.from_function(
            func=calculator,
            name="calculator",
            description="安全计算基础算术表达式，不执行 Python 代码。",
            args_schema=CalculatorInput,
        ),
        StructuredTool.from_function(
            func=data_query,
            name="data_query",
            description="按工作表和包含条件查询企业 Excel 数据，返回结构化行。",
            args_schema=DataQueryInput,
        ),
        StructuredTool.from_function(
            func=web_search,
            name="web_search",
            description="通过 Tavily 搜索公开网页；需要配置 TAVILY_API_KEY。",
            args_schema=WebSearchInput,
        ),
        StructuredTool.from_function(
            func=document_analysis,
            name="document_analysis",
            description="读取指定企业文档并提取结构、长度和与问题相关的段落。",
            args_schema=DocumentAnalysisInput,
        ),
    ]
